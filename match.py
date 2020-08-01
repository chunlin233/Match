import pandas as pd
import numpy as np
import os
# import xlwt
import tkinter as tk
from tkinter import *
from tkinter import filedialog
import tkinter.messagebox
import openpyxl
import xlrd


from openpyxl.styles import Font, Border, Side, PatternFill, colors, Alignment
from openpyxl import load_workbook, Workbook
# from openpyxl.styles import Font #导入字体模块
#
# from openpyxl.styles import PatternFill #导入填充模

global selectFile
global downloadfile
global matched_file_path
global row
global list_emails_all
global short_list, full_list, re_email, paper_list, pd_list, pb_list, pi_list, re_list, affiliation_list1, affiliation_list2, email, author
global paper_index, RGB_Code
short_list = []
full_list = []
re_email = []
paper_list = []
pd_list = []
pb_list = []
pi_list = []
re_list = []
affiliation_list1 = []
affiliation_list2 = []
email = []
author = []
rp = []

# 匹配度阈值，若低于该值，则匹配输出为None
score_cutoff = 60
#global email_split

###########################分开一篇文章中的多个作者名##################################################
def name_array_split(one_paper_name):
    name_array = one_paper_name.split('; ')
    return name_array


###########################分开一片文章中的多个通讯邮箱#################################################
def email_array_split(one_paper_email):
    email_array = one_paper_email.split('; ')
    return email_array

############################寻找后缀为163的邮箱########################################################
def email_find_163(email_array, list_all_emails):
    # global row
    # index=0
    match_list = []
    row_list = []
    for j in range(len(re_email)):
        if re_email[j].find('@gmail.com') > 0:
            # index = index + 1
            match_list.append(re_email[j])
            row_list.append(j)
    return row_list


##############################分割国家####################################
def differ_country(rp_names_array,email_array):
    country = []
    last_country_str = []
    last_country = []

    # for i in range(len(rp_names_array)):
    #     if pd.isna(rp_names_array[i]):
    #         continue
    for i in range(len(email_array)):
        if pd.isna(email_array[i]): #设定在email为空的情况下不会输出
            last_country.append('NA')
            continue
        country.append(rp_names_array[i])
        last = rp_names_array[i].split(', ')
        last_str = last[len(last)-1].split('.')[0]
        # print(type(last_str))
        # print(last_str)
        last1 = last_str.split(' ')
        last_country.append(last1[len(last1)-1])

    # print(last_country)
    # print(len(last_country))
    return last_country



#############################读取国家表格内容并对不同州赋予不同的颜色#######################################
def open_country():
    excel_country = []  #将国家信息和大洲信息存储到数组中
    excel_state = []
    rcountry = xlrd.open_workbook('data/country.xlsx')
    rcountry.sheets()
    rsheet = rcountry.sheet_by_index(0)

    # 循环工作簿的所有行
    for row1 in rsheet.get_rows():
        country_column = row1[1]  # 英文国家名所在的列
        country_value = country_column.value  # 国家名
        if country_value != '英文名称':  # 排除第一行
            excel_country.append(country_value)
            # print("英文country", country_value)
    for row2 in rsheet.get_rows():
        state_column = row2[3]  # 英文国家名所在的列
        state_value = state_column.value  # 州名
        if state_value != '所属洲':  # 排除第一行
            excel_state.append(state_value)

    # print(excel_state)
    # print(len(excel_state))
    color_list = [] #country中颜色的十六进制
    #将按照顺序列表的颜色存入数组
    for i in range(len(excel_state)):
        if excel_state[i] == '亚洲':
            RGB_Code = 'FFC0CB'
        elif excel_state[i] == '欧洲':
            RGB_Code = 'DC143C'
        elif excel_state[i] == '大洋洲':
            RGB_Code = '00FFFF'
        elif excel_state[i] == '非洲':
            RGB_Code = '008B8B'
        elif excel_state[i] == '北美洲':
            RGB_Code = 'FFD700'
        elif excel_state[i] == '南美洲':
            RGB_Code = 'FFDEAD'
        else:
            RGB_Code = 'FFFFFF'
        color_list.append(RGB_Code)
    return color_list,excel_country,excel_state


####################################匹配国家和州，并获取匹配上的行数#######################################################
def match_country(last_country, excel_country, excel_state, color_list, email_array):
    state = []
    color = []
    count1 = 0
    row_index = []

    for k in range(len(email_array)):
        flag = 0
        if pd.isna(email_array[k]):
            count1 = count1 + 1
            continue
        for j in range(len(excel_country)):

            if last_country[k] == excel_country[j]:
                flag = 1
                # index = k+count1
                # row_index.append(index)
                cur_color = color_list[j]
                emails_one = email_array_split(email_array[k])
                for z in range(len(emails_one)):
                    color.append(cur_color)
                continue
        if flag == 0:
            emails_one = email_array_split(email_array[k])
            for a in range(len(emails_one)):
                color.append('FFFFFF')
    # print(len(email_array)) #71
    #
    # print(len(last_country))    #71
    # print(len(row_index))
    # print(color)
    # print(len(color))
    return color


#######################################换单元格颜色##############################################
def changecolor(row_list, color):
    # mf.save(matched_file_path)
    global matched_file_path
    y=[]
    mf = openpyxl.load_workbook(matched_file_path)  # 加载已经存在的excel
    mf_name = mf.sheetnames
    # print(mf_name)
    fille = PatternFill('solid', fgColor='FFFFFF')  # 设置填充颜色为 橙色
    mf_sheet = mf[mf_name[0]]

    for i in range(len(color)):
        # x=i+2
        mf_sheet['B%d' %(i+2)].fill = PatternFill('solid', fgColor='%s' %color[i])
    mf.save(matched_file_path)



########################################简称匹配全称###############################################
def match_full_short_name(short_names, full_names, rp_names):
    import process
    import fuzz

    def fuzz_match(short_names, full_names, rp_names):
        matched_full_names = []
        for i in range(len(rp_names)):
            # name = process.extractOne(rp_names[i], names, scorer=fuzz.partial_token_set_ratio)
            # print(names)
            max_ratio = 0
            max_ratio_index = -1
            for j in range(len(short_names)):
                # name = process.extractOne(rp_names[i], names, scorer=fuzz.partial_token_set_ratio)
                ratio = fuzz.partial_ratio(rp_names[i], short_names[j])
                if max_ratio < ratio:
                    max_ratio = ratio
                    max_ratio_index = j
            matched_full_names.append(full_names[max_ratio_index])

        return matched_full_names

    if type(rp_names) != list:
        rp_names = [rp_names]
    if type(short_names) != list:
        short_names = [short_names]
    if type(full_names) != list:
        full_names = [full_names]
    matched_full_names = fuzz_match(short_names, full_names, rp_names)
    # print(rp_names) #缩写的匹配上的通讯作者名
    return matched_full_names


#########################################邮箱匹配名字################################################
def match_name_email(names, emails):
    import process
    import fuzz

    def fuzz_match(names, emails):
        # print(names)
        matched_short_names = []
        matched_full_names = []
        for i in range(len(emails)):
            simple_email = emails[i].split('@')[0].strip()
            simple_email = re.sub(r'[_.]', ' ', simple_email)
            # name = process.extractOne(simple_email, names, scorer=fuzz.partial_token_set_ratio)   #changed
            """ changed"""
            name = process.extractOne(simple_email, names, scorer=fuzz.partial_token_set_ratio,score_cutoff=score_cutoff)
            if name is None:
                matched_short_names.append('NA')
                matched_full_names.append('NA')
            else:
                short, full = name[0].split('; ')
                matched_short_names.append(short)
                matched_full_names.append(full)
            """ changed"""

        return matched_short_names, matched_full_names

    if type(names) != list:
        names = [names]
    if type(emails) != list:
        emails = [emails]
    lists = fuzz_match(names, emails)
    return lists


#########################################邮箱匹配rp#####################################################
""" add """
############################
# 邮箱匹配RP，输出匹配的RP作者简写
def match_rp_email(rp_names, emails):
    import process
    import fuzz

    def fuzz_match(rp_names, emails):
        # print(names)
        matched_rp_names = []
        for i in range(len(emails)):
            simple_email = emails[i].split('@')[0].strip()
            simple_email = re.sub(r'[_.]', ' ', simple_email)
            name = process.extractOne(simple_email, rp_names, scorer=fuzz.partial_token_set_ratio, score_cutoff=score_cutoff)
            if name is None:
                matched_rp_names.append("NA")
            else:
                matched_rp_names.append(name[0])

        return matched_rp_names

    if type(rp_names) != list:
        rp_names = [rp_names]
    if type(emails) != list:
        emails = [emails]
    matched_rp_names = fuzz_match(rp_names, emails)
    return matched_rp_names
""" add """


##########################################带入邮箱、论文名、出版时间、出版商信息、城市信息、机构信息1、机构信息2#####################
def message(email_array, list_paper, list_date, list_publisher, list_city, list_reference, reprint_author, full_names):
    count1 = 0
    count2 = 0
    count3 = 0
    count4 = 0
    count5 = 0
    count6 = 0
    count7 = 0
    count8 = 0
    count9 = 0
    length = len(email_array)
    list_all_emails = []


    ##################################################
    # email = []  #将所有的email带入
    for h in range(len(email_array)):
        if pd.isna(email_array[h]):
            count8 = count8 + 1
            continue
        emails_one = email_array_split(email_array[h])
        for t in range(len(emails_one)):
            each_email = emails_one[t]
            re_email.append(each_email)


    ##################################################
    paper_name = []  # 将所有的论文名称存入了paper_name
    for j in range(len(list_paper)):

        if pd.isna(email_array[j]):
            x = j - count1
            count1 = count1 + 1
            # count_list.append(x)
            continue
        paper_title = list_paper[j]
        paper_name.append(paper_title)
        emails_one = email_array_split(email_array[j])

        each_paper = paper_name[j - count1]
        for a in range(len(emails_one)):
            paper_list.append(each_paper)

    ##################################################
    date = []  # 将论文出版时间带入
    for k in range(len(list_date)):
        if pd.isna(email_array[k]):
            # x = j - count
            count2 = count2 + 1
            continue
        each_date = list_date[k]
        date.append(each_date)
        emails_one = email_array_split(email_array[k])
        # each_email = date[k - count2]
        for b in range(len(emails_one)):
            pd_list.append(date[k - count2])


    #################################################
    publisher = []  # 将杂志信息带入
    for p in range(len(list_publisher)):
        if pd.isna(email_array[p]):
            count3 = count3 + 1
            continue
        each_publisher = list_publisher[p]
        publisher.append(each_publisher)
        emails_one = email_array_split(email_array[p])
        for c in range(len(emails_one)):
            pb_list.append(publisher[p - count3])

    ##################################################
    country = []  # 将城市信息带入
    last_country = []
    for v in range(len(email_array)):
        if pd.isna(email_array[v]):
            count4 = count4 + 1
            continue
        country.append(reprint_author[v])
        last = reprint_author[v].split(', ')
        last_str = last[len(last) - 1].split('.')[0]
        last1 = last_str.split(' ')
        last_country.append(last1[len(last1) - 1])
        emails_one = email_array_split(email_array[v])
        for z in range(len(emails_one)):
            pi_list.append(last_country[v - count4])
        # each_city = list_city[v]
        # city.append(each_city)
        # emails_one = email_array_split(email_array[v])
        # for z in range(len(emails_one)):
        #     pi_list.append(city[v - count4])

    #################################################
    reference = []  # 将引用次数信息带入
    for m in range(len(list_reference)):
        if pd.isna(email_array[m]):
            count5 = count5 + 1
            continue
        each_reference = list_reference[m]
        reference.append(each_reference)
        emails_one = email_array_split(email_array[m])
        for q in range(len(emails_one)):
            re_list.append(reference[m - count5])

    #####################################################
    affiliation1 = []   #将RP一列的机构信息带入
    for o in range(len(reprint_author)):
        if pd.isna(email_array[o]):
            count6 = count6 + 1
            continue
        each_affiliation = reprint_author[o]
        affiliation1.append(each_affiliation)
        emails_one = email_array_split(email_array[o])
        for z in range(len(emails_one)):
            affiliation_list1.append(affiliation1[o - count6])

    #######################################################
    affiliation2 = []   #将RP中提取出的机构名称带入
    for s in range(len(reprint_author)):
        if pd.isna(email_array[s]):
            count7 = count7+1
            continue
        each_affiliation1= reprint_author[s].split(', ')[2]
        affiliation2.append(each_affiliation1)
        emails_one = email_array_split(email_array[s])
        for f in range(len(emails_one)):
            affiliation_list2.append(affiliation2[s-count7])

    ##########################################################
    full =[]
    all_author = []
    all_rp = []

    # for r in range(len(email_array)):
    #     if pd.isna(email_array[r]):
    #         count9 = count9 + 1
    #         continue
    #     each_full_author = full_names[r]
    #     full.append(each_full_author)
    #     emails_one = email_array_split(email_array[r])
    #     for b in range(len(emails_one)):
    #         all_author.append(full[r-count9])

    for i in range(len(email_array)):
        if pd.isna(email_array[i]):
            continue
        emails_one = email_array_split(email_array[i])
        for _ in range(len(emails_one)):
            all_author.append(full_names[i])
            all_rp.append(reprint_author[i])

    # print(all_author)
    # print(len(all_author))

    for u in range(len(all_author)):
        if full_list[u] == 'NA':
            author.append(all_author[u])
            rp.append(all_rp[u])
        else:
            author.append('')
            rp.append('')
    # print(author)
    print(len(author))
        # emails_one = email_array_split(email_array[r])
        # if full_list[r] != 'NA':
        #     for y in range(len(emails_one)):
        #         author.append(full_list[r])
        # else:
        #     # emails_one = email_array_split(email_array[r])
        #     for t in range(len(emails_one)):
        #         author.append('')
        # print(full_names)

    # for u in range(len(short_list)):
        #     if full_list[u] != 'NA':
        #         author.append('')
        #     else:
        #         emails_one = email_array_split(email_array[r])
        #         for y in range(len(emails_one)):
        #             author.append(full_list[r-count9])
    # print(author)
    # print(len(author))

    ###########################################################
    for x in range(length):
        # print(x)
        # print(email_array[x])
        list_all_emails.append(email_array[x])

    return list_all_emails



##########################################################################################################
def match_mul_paper(short_name_array, full_name_array, rp_name_array, email_array):
    global short_list
    global full_list
    length = len(email_array)
    new_emails = []

    for i in range(length):
        if pd.isna(email_array[i]):
            continue
        emails = email_array_split(email_array[i])
        [new_emails.append(e) for e in emails]

        emails = email_array_split(email_array[i])
        short_names = name_array_split(short_name_array[i])
        full_names = name_array_split(full_name_array[i])
        names = ['; '.join([short, full]) for short, full in zip(short_names, full_names)]
        unsplited_rp_names = re.findall(r'[; ]?(.*?)\s\(reprint author\).*?\.', rp_name_array[i])
        rp_names = []
        for rp_name in unsplited_rp_names:
            rp_name = rp_name.split(';')
            [rp_names.append(r.strip()) for r in rp_name]

        if pd.isna(rp_name_array[i]) or len(rp_names) == 0:
            # 若rp无数据，则用邮箱匹配names
            matched_short_names, matched_full_names = match_name_email(names, emails)
            [short_list.append(s.split(',')[0]) for s in matched_short_names]
            [full_list.append(s) for s in matched_full_names]
            assert len(matched_short_names) == len(emails)
            assert len(matched_full_names) == len(emails)


        elif len(emails) >= 1:
            # 若rp有数据(>=1)，则先用邮箱匹配rp
            matched_rp_names = match_rp_email(rp_names, emails)
            assert len(matched_rp_names) == len(emails)

            matched_full_names = ['NA'] * len(emails)

            for i in range(len(matched_rp_names)):
                if matched_rp_names[i] != 'NA':
                    # 邮箱匹配rp成功，简称匹配全称
                    matched_full_name = match_full_short_name(short_names, full_names, matched_rp_names[i])
                    matched_full_names[i] = matched_full_name[0]

                elif matched_rp_names[i] == 'NA':
                    # 若邮箱匹配rp失败，则尝试邮箱匹配names
                    matched_short_name, matched_full_name = match_name_email(names, emails[i])
                    matched_rp_names[i] = matched_short_name[0]
                    matched_full_names[i] = matched_full_name[0]

            [short_list.append(s.split(',')[0]) for s in matched_rp_names]
            [full_list.append(s) for s in matched_full_names]

            assert len(matched_rp_names) == len(emails)
            assert len(matched_full_names) == len(emails)

    return new_emails


###################################写入新的EXCEL文件#########################################
def write_to_excel(matched_file_path):
    global selectFile
    global short_list
    global re_email
    global paper_list
    global pd_list



    dr = pd.read_excel(selectFile)

    df = pd.DataFrame()
    df.insert(df.shape[1], 'Full_Name', full_list)
    df.insert(df.shape[1], 'First_Name', short_list)
    df.insert(df.shape[1], 'Paper_Title', paper_list)
    df.insert(df.shape[1], 'Emails', re_email)
    df.insert(df.shape[1], 'Country', pi_list)
    df.insert(df.shape[1], 'JOURNAL', pb_list)
    df.insert(df.shape[1], 'Publish_Year', pd_list)
    df.insert(df.shape[1], 'Reference_Count', re_list)
    df.insert(df.shape[1], 'Affiliation1', affiliation_list1)
    df.insert(df.shape[1], 'Affiliation2', affiliation_list2)
    df.insert(df.shape[1], 'Author', author)
    df.insert(df.shape[1], 'RP', rp)

    # df['EAS'] = df_add['EAS']
    # df['EAF'] = df_add['EAF']
    df.to_excel(matched_file_path)




###################################执行匹配的函数（不是匹配算法）############################################
# def match(file_path, matched_file_path):
def match():
    global selectFile
    global downloadfile
    global matched_file_path
    print(selectFile)

    file_path = selectFile
    xls = pd.ExcelFile(file_path)
    filename = os.path.split(file_path)[-1].split('.')[0]
    matched_file_path = os.path.join(downloadfile+'/' + filename + '_matched.xlsx')

    data = pd.read_excel(xls, "Sheet1")
    short_names = data['AU']
    full_names = data['AF']
    reprint_authers = data['RP']
    author_emails = data['EM']
    paper_title = data['TI']
    paper_date = data['PY']
    paper_publisher = data['SO']
    paper_city = data['PI']
    paper_reference = data['NR']



    email_array = match_mul_paper(short_names, full_names, reprint_authers, author_emails)
    list_all_emails = message(author_emails, paper_title, paper_date, paper_publisher, paper_city, paper_reference, reprint_authers, full_names)
    last_country = differ_country(reprint_authers, author_emails)
    color = match_country(last_country,excel_country, excel_state, color_list, author_emails)

    row_list = email_find_163(email_array,list_all_emails)


    # print(short_list)
    # print(len(short_list))
    # print(full_list)
    # print(len(full_list))
    # print(full_list)


    # emails=email_array_split(author_emails)
    #
    # email_find_163(emails)



    # emails=email_array_split(author_emails)
    # email_find_163(author_emails)
    # print(author_emails)

    write_to_excel(matched_file_path)
    changecolor(row_list,color)
    # email_find_163()

    xls.close()
    tk.messagebox.showinfo('通讯作者匹配系统', '匹配已完成！')

##############################################################################
#设计文件上传
def upload():
    global selectFile
    selectFile = tk.filedialog.askopenfilename()  # askopenfilename 1次上传1个；askopenfilenames1次上传多个
    e1.insert(0, selectFile)
    print(selectFile)
    return selectFile

#设计结果下载
def select():
    global downloadfile
    downloadfile=tk.filedialog.askdirectory()
    e2.insert(0, downloadfile)
    print(downloadfile)

# def download():
#     global downloadfile
#
#     print("download")



#退出
def exit():
    window.destroy()


#########################################################

if __name__ == "__main__":

    ####################################################################################################################
    # file_path = "./data/2d.xlsx"
    # file_path = selectFile
    # filename = os.path.split(file_path)[-1].split('.')[0]
    # matched_file_path = os.path.join('./data', filename + '_matched.xlsx')
    # print("processing...")
    # try:
    #     match(file_path, matched_file_path)
    #     print("successful")
    # except Exception as e:
    #     print(e)

    #################################################tk界面设计##########################################################
    # 第1步，实例化object，建立窗口window
    window = tk.Tk()

    # 第2步，给窗口的可视化起名字
    window.title('通讯作者匹配系统')

    # 第3步，设定窗口的大小(长 * 宽)
    window.geometry('800x400')  # 这里的乘是小x



    # 第4步，在图形界面上设定并放置标签
    l = tk.Label(window, text='通讯作者匹配系统', bg='green', font=('Arial', 12), width=30, height=2)
    # 说明： bg为背景，font为字体，width为长，height为高，这里的长和高是字符的长和高，比如height=2,就是标签有2个字符这么高

    l.pack()  # Label内容content区域放置位置，自动调节尺寸
    # 放置lable的方法有：1）l.pack(); 2)l.place();

    # 第5步，在窗口界面设置放置Button按键
    b1 = tk.Button(window, text='上传文件', font=('Arial', 12), width=10, height=1, command=upload)

    b5 = tk.Button(window, text='选择下载路径', font=('Arial', 12), width=10, height=1, command=select)
    # 在图形界面上设定输入框控件entry并放置控件
    e1 = tk.Entry(window, show=None, font=('Arial', 12), width=50)  # 显示成明文
    e2 = tk.Entry(window, show=None, font=('Arial', 12), width=50)  # 显示成明文

    b4 = tk.Button(window, text='匹配并导出', font=('Arial', 12), width=10, height=1, command=match)
    b1.place(x=100, y=50)
    b5.place(x=100, y=100)
    e1.place(x=230, y=55)
    e2.place(x=230, y=105)


    b4.place(x=600, y=250)

    l2 = tk.Label(window, text='使用说明：\n1.将需要匹配的文件导入到系统中\n2.选择导出路径\n3.点击匹配下载按钮可以在指定路径拥有匹配信息excel文件',
                  font=('Arial', 12), height=4, fg='red')
    l2.place(x=250, y=320)

    b3 = tk.Button(window, text='返回', font=('Arial', 12), width=10, height=1, command=exit)
    b3.place(x=600, y=300)

    # b2 = tk.Button(window, text='换背景色', font=('Arial', 12), width=10, height=1)
    # b2.place(x=600, y=350)
    ############################################

    # matched_file_path=None
    # file_path = selectFile
    #print(file_path)

    # if file_path != None:
    # filename = os.path.split(file_path)[-1].split('.')[0]
    # matched_file_path = os.path.join('./data', filename + '_matched.xlsx')
    print("processing...")
    try:
        # match(file_path, matched_file_path)
        color_list, excel_country, excel_state = open_country()
        match()

        print("successful")
    except Exception as e:
        print(e)
    ############################################
    # 第6步，主窗口循环显示
    window.mainloop()
    # 注意，loop因为是循环的意思，window.mainloop就会让window不断的刷新，如果没有mainloop,就是一个静态的window,传入进去的值就不会有循环，mainloop就相当于一个很大的while循环，有个while，每点击一次就会更新一次，所以我们必须要有循环
    # 所有的窗口文件都必须有类似的mainloop函数，mainloop是窗口文件的关键的关键。


    ####################################################################################################################
